import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import os

# Setup WebDriver
driver = webdriver.Chrome()

# Read host links from CSV file
host_info_df = pd.read_csv('D:/host_info.csv')

# File path to save the scraped data
output_file = 'D:/scraped_host_info.xlsx'

# Check if the output file exists
file_exists = os.path.isfile(output_file)

# Open the workbook if it exists, else create a new one
if file_exists:
    workbook = load_workbook(output_file)
else:
    workbook = None

# Use ExcelWriter to manage writing data to the workbook
with pd.ExcelWriter(output_file, engine='openpyxl', mode='a' if file_exists else 'w') as writer:
    if file_exists:
        writer.book = workbook

    # If the file does not exist, write the header
    if not file_exists:
        header_df = pd.DataFrame(columns=['hostlink', 'name', 'description', 'total_reviews', 'rating', 'years_hosting', 'listing_count'])
        header_df.to_excel(writer, sheet_name='Sheet1', index=False)

    # Track the number of processed hostlinks
    processed_count = 0
    total_count = len(host_info_df)

    # Iterate through each link in the hostlink column
    for index, row in host_info_df.iterrows():
        hostlink = row['hostlink']
        processed_count += 1

        print(f"Processing {processed_count}/{total_count}: {hostlink}")

        try:
            # Open the webpage
            driver.get(hostlink)

            # Wait for the page to fully load
            wait = WebDriverWait(driver, 10)  # 10 seconds timeout

            # Initialize variables with default values
            name = ''
            description = ''
            total_reviews = ''
            rating = ''
            years_hosting = ''
            listing_count = ''

            # Scrape the required details with error handling
            try:
                name_element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'span.swe026z')))
                name = name_element.text if name_element else ''
            except Exception as e:
                print(f"Error scraping name: {e}")

            try:
                description_element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'span._1e2prbn')))
                description = description_element.text if description_element else ''
            except Exception as e:
                print(f"Error scraping description: {e}")

            try:
                total_reviews_element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'span[data-testid="Reviews-stat-heading"]')))
                total_reviews = total_reviews_element.text if total_reviews_element else ''
            except Exception as e:
                print(f"Error scraping total reviews: {e}")

            try:
                rating_element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div.ruujrrq')))
                rating = rating_element.text if rating_element else ''
            except Exception as e:
                print(f"Error scraping rating: {e}")

            try:
                years_hosting_element = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'span[data-testid="Years hosting-stat-heading"]')))
                years_hosting = years_hosting_element.text if years_hosting_element else ''
            except Exception as e:
                print(f"Error scraping years hosting: {e}")

            try:
                button = wait.until(EC.visibility_of_element_located((By.XPATH, '//button[contains(text(), "View all")]')))
                listing_count = button.text
            except Exception as e:
                print(f"Error scraping listing count: {e}")
                listing_count = 'Less than 10 listings'

            # Prepare the scraped data
            data = {
                'hostlink': hostlink,
                'name': name,
                'description': description,
                'total_reviews': total_reviews,
                'rating': rating,
                'years_hosting': years_hosting,
                'listing_count': listing_count
            }

            # Convert the data to a DataFrame
            df = pd.DataFrame([data])

            # Append the scraped data to the Excel file incrementally
            df.to_excel(writer, sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, header=False, index=False)

        except Exception as e:
            print(f"An error occurred for {hostlink}: {e}")
            continue

# Close the browser
driver.quit()

print(f"Scraping completed and data saved to '{output_file}'")
